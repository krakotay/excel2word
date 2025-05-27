from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from copy import copy
from time import time

SAVE_PATH = Path("output")
SAVE_PATH.mkdir(exist_ok=True)

def copy_sheet(src_ws: Worksheet, dst_ws: Worksheet):
    # Копируем значения и стили
    for row in src_ws.iter_rows():
        for cell in row:
            row_idx = cell.row
            # для обычных ячеек берем col_idx, для merged — преобразуем letter -> index
            if hasattr(cell, 'col_idx'):
                col_idx = cell.col_idx
            else:
                col = cell.column  # Может быть буквой
                col_idx = col if isinstance(col, int) else column_index_from_string(col)
            new_cell = dst_ws.cell(row=row_idx, column=col_idx, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                # Прямо присваиваем строковый код формата
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    # Воссоздаём merged ranges
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
    # Копируем размеры строк
    for idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[idx].height = dim.height
    # Копируем ширины столбцов
    for idx, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[idx].width = dim.width
    # Копируем остальные свойства листа
    dst_ws.sheet_format = src_ws.sheet_format
    dst_ws.sheet_properties = src_ws.sheet_properties
    dst_ws.page_margins = src_ws.page_margins
    dst_ws.page_setup = src_ws.page_setup
    dst_ws.print_options = src_ws.print_options

    # --- Новый блок: копируем условное форматирование ---
    # Простой (хлороформатный) способ — скопировать «правила» целиком:
    dst_ws.conditional_formatting._cf_rules = copy(src_ws.conditional_formatting._cf_rules)

# Load workbooks
ORIGIN_PATH = Path('шаблоны', 'task2')
origin = ORIGIN_PATH / 'origin.xlsx'
target = ORIGIN_PATH / 'пустой.xlsx'
origin_wb = load_workbook(origin)
target_wb = load_workbook(target)

for sheet_name in reversed(origin_wb.sheetnames):
    # Удаляем существующий лист, если он есть
    if sheet_name in target_wb.sheetnames:
        target_wb.remove(target_wb[sheet_name])
    # Создаём новый лист в позиции 0 (в начале)
    dst = target_wb.create_sheet(title=sheet_name, index=0)
    src = origin_wb[sheet_name]
    copy_sheet(src, dst)

# Optionally remove the default 'Sheet' if it's empty and not in origin
if 'Sheet' in target_wb.sheetnames and 'Sheet' not in origin_wb.sheetnames:
    target_wb.remove(target_wb['Sheet'])

# Save changes
timestamp = str(int(time()))
tmp = SAVE_PATH / timestamp
tmp.mkdir()

target_wb.save(tmp / f'target_{target.name}')
