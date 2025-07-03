from pathlib import Path
import openpyxl
import polars as pl
from tqdm import tqdm

def extract_spec_tables(xlsx_path: Path) -> list[pl.DataFrame]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True, keep_links=False)
    tables: list[pl.DataFrame] = []

    for sheet_name in tqdm(wb.sheetnames):
        ws = wb[sheet_name]

        # Пройдемся по всем строкам, чтобы найти заголовки таблиц
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if val and isinstance(val, str) and val.startswith("spec_"):
                    start_row = cell.row
                    start_col = cell.column  # это номер, не буква

                    # Считываем имена 4 столбцов из строки-заголовка
                    headers = []
                    for offset in range(4):
                        hdr_cell = ws.cell(row=start_row, column=start_col + offset)
                        headers.append(str(hdr_cell.value or "").strip())

                    # Собираем данные до первого полностью пустого ряда
                    data_rows = []
                    curr_row = start_row + 1
                    while True:
                        row_vals = []
                        # забираем все 4 колонки
                        for offset in range(4):
                            v = ws.cell(row=curr_row, column=start_col + offset).value
                            row_vals.append(v)
                        # если все четыре ячейки пустые — это конец таблицы
                        if all(v is None for v in row_vals):
                            break
                        data_rows.append(row_vals)
                        curr_row += 1

                    # Если под заголовком были данные, делаем DataFrame
                    if data_rows:
                        # соберём dict из списков по колонкам
                        cols = {headers[i]: [str(row[i]) for row in data_rows] for i in range(4)}
                        df = pl.DataFrame(cols)
                        df = df.filter((pl.nth(-1) != "0") & (pl.nth(-2) != "0"))
                        short_df = df.select(pl.nth(-1), pl.nth(-2)).with_columns(pl.all().replace("None", None))
                        short_df = short_df[[s.name for s in short_df if not (s.null_count() == short_df.height)]]
                        # по желанию можно добавить колонку с именем листа
                        if not short_df.is_empty() and not df.is_empty():
                            tables.append(df.with_columns(pl.lit(sheet_name).alias("__sheetname__")))

                    # пропускаем остальные ячейки в этой же строке,
                    # чтобы не дублировать одну и ту же таблицу
                    break

    return tables

if __name__ == "__main__":
    ORIGIN_PATH = Path("шаблоны/task6")  # Укажите ваш путь
    path = ORIGIN_PATH / 'Отчет-МСФО-6-2024-Индо.xlsx'
    all_tables = extract_spec_tables(path)
    # Пример: вывести, сколько таблиц нашли
    print(f"Найдено таблиц: {len(all_tables)}")
    # Например, посмотреть первые 5 строк первой таблицы
    if all_tables:
        for table in all_tables:

            print(table.head())
        # print(all_tables[0].head())
    print(f"Найдено таблиц: {len(all_tables)}")
